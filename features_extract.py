# ------------------------------------------------------------------------------------------------------------------
# 特征提取
import numpy as np
import pandas as pd
import os
import SimpleITK as sitk

col_y = "FEV1_MEAS" #TODO: 更改预测目标-列名, FEV1_MEAS, FVC_MEAS
PATH_data ="G:/crop_padded" #TODO: 更改ct图像（.hdr）索引路径
PATH_excel ="E:/pft_data_DL.xlsx" #TODO：更改 数据集.xlsx 所在位置
weight_num = '092' #TODO: check
os.chdir("E:/training") #TODO: 更改当前工作目录-training
print('当前工作目录：', os.getcwd())

img_width, img_height = 180,178 # 图像大小，暂不调整
NUM_CH = 1 # 灰度图像，通道数为1，无需更改
NUM_FRAMES = 140 #TODO：更改每ct系列导入图像张数

# 导入数据集
df_info = pd.read_excel(PATH_excel,
                        index_col = 0, # 将第一列作为索引，‘NO’列
                       engine = "openpyxl")

list_new_key = [] # 传输列名
for _key in df_info.keys():
    _new_key = _key.strip()
    list_new_key.append(_new_key)
df_info.columns = list_new_key
df_info = df_info[df_info[col_y].notnull()] # 检查col_y是否非空，为空的行不被导入df_info中

def get_path_in_workspace(_row):
    # _group_path = _row["group_path"]
    _ID = _row["examID"]
    # _path_in_workspace = os.path.join(PATH_data, _group_path, str(_ID) + ".nii.hdr")
    _path_in_workspace = os.path.join(PATH_data, str(_ID) + ".nii.img")
    return _path_in_workspace

df_train = df_info[df_info["set_split"] =="train"].reset_index(drop = True)
df_valid = df_info[df_info["set_split"] =="valid"].reset_index(drop = True)
df_test = df_info[(df_info["set_split"] == "test")].reset_index(drop=True)

# 导入ct图像
id_train = list(range(len(df_train)))  # 载入索引
id_valid = list(range(len(df_valid)))
id_test = list(range(len(df_test)))

# 图像预处理方法
def preprocess_12bit(input_img):  # 12 bit full range
    density_low = -1024 #TODO：可调整
    density_high = 3071 #TODO：可调整

    output_img = (input_img - density_low) / (density_high - density_low + 1)
    output_img[output_img < 0.] = 0.
    output_img[output_img > 1.] = 1.

    return output_img

# 加载图像
def load_image(path,augmentation = False, num_channel = NUM_CH):
    img = sitk.GetArrayFromImage(sitk.ReadImage(path))

    if num_channel == 1: # 灰度图
        img_12bit = preprocess_12bit(img)  # (z,y,x) # 图像预处理方法1

        img_tmp = img_12bit
        img_tmp = img_tmp[:, ::-1, :]  # (z,-y,x)
        img_tmp = np.expand_dims(img_tmp, axis=-1)
        import scipy.ndimage
        img_resize = scipy.ndimage.zoom(img_tmp, zoom=
                                  (NUM_FRAMES/img_tmp.shape[0], # z
                                  img_height/img_tmp.shape[1], # y
                                  img_width/img_tmp.shape[2], # x
                                   NUM_CH),order=3)
        img_out = img_resize.copy()


    return img_out



from i3d_inception_last_global import Inception_Inflated3d as I3d
#------
from keras.models import Model
from keras.layers import Dense, Input, concatenate
from keras.utils.vis_utils import plot_model
from keras import backend as K
K.clear_session()

n_classes = 1

# 用全局平均池化更改最后一部分
base_model = I3d(include_top=False,
                weights=None,
#                 input_tensor=None,
                input_shape=(NUM_FRAMES,img_height,img_width,NUM_CH),
#                 dropout_prob=0.0,
                endpoint_logit=False, # softmax will be applied
                classes=n_classes
                )
#----------------------------------
x = base_model.output
x = Dense(500, activation = "relu")(x)
prediction = Dense(1,activation = "linear" )(x)
model = Model(inputs=base_model.input, outputs=prediction)
model.summary()

# 创建用于提取特征的新模型
feature_extraction_model = Model(inputs=model.input, outputs=model.get_layer('dense').output) # 倒数第二层

# 加载模型权重
model.load_weights(f'train_weights_epoch_{weight_num}.h5')

# 提取特征函数
def extract_features(data):
    features = feature_extraction_model.predict(data, verbose=1)
    return features

import openpyxl
from openpyxl import Workbook
train_features_path = 'train_features.xlsx'
valid_features_path = 'valid_features.xlsx'
test_features_path = 'test_features.xlsx'

batch_size = 1
# # #
for i in range(len(df_valid)):
    valid_img = np.zeros((batch_size, NUM_FRAMES, 178, 180, NUM_CH))
    for j in range(batch_size):
            _idx = id_valid[i * batch_size + j]
            print(_idx)
            _filename = df_valid.loc[_idx, "path_in_workspace"]
            valid_img_t = load_image(_filename, augmentation=False)
            valid_img[j] = valid_img_t
            valid_features_t = extract_features(valid_img)
            valid_df_t = pd.DataFrame(valid_features_t)
            valid_data_m = valid_df_t.iloc[0].tolist()
            blank = ' '
            valid_data = blank.join('%s' % a for a in valid_data_m)
            valid_features = openpyxl.load_workbook(valid_features_path)
            valid_features_s = valid_features.active
            sheetnames = valid_features.get_sheet_names()
            table = valid_features.get_sheet_by_name(sheetnames[0])
            table = valid_features.active
            nrows = valid_features_s.max_row  # 获得行数
            ncolumns = table.max_column  # 获得行数
            values = [valid_data]
            for value in values:
                table.cell(nrows + 1, 1).value = value
                nrows = nrows + 1
            valid_features.save(valid_features_path)


for i in range(len(df_test)):
    test_img = np.zeros((batch_size, NUM_FRAMES, 178, 180, NUM_CH))
    for j in range(batch_size):
            _idx = id_test[i * batch_size + j]
            print(_idx)
            _filename = df_test.loc[_idx, "path_in_workspace"]
            test_img_t = load_image(_filename, augmentation=False)
            test_img[j] = test_img_t
            test_features_t = extract_features(test_img)
            test_df_t = pd.DataFrame(test_features_t)
            test_data_m = test_df_t.iloc[0].tolist()
            blank = ' '
            test_data = blank.join('%s' % a for a in test_data_m)
            test_features = openpyxl.load_workbook(test_features_path)
            test_features_s = test_features.active
            sheetnames = test_features.get_sheet_names()
            table = test_features.get_sheet_by_name(sheetnames[0])
            table = test_features.active
            nrows = test_features_s.max_row  # 获得行数
            ncolumns = table.max_column  # 获得行数
            values = [test_data]
            for value in values:
                table.cell(nrows + 1, 1).value = value
                nrows = nrows + 1
            test_features.save(test_features_path)

#
for i in range(len(df_train)):
    train_img = np.zeros((batch_size, NUM_FRAMES, 178, 180, NUM_CH))
    for j in range(batch_size):
            _idx = id_train[i * batch_size + j]
            print(_idx)
            _filename = df_train.loc[_idx, "path_in_workspace"]
            train_img_t = load_image(_filename, augmentation=False)
            train_img[j] = train_img_t
            train_features_t = extract_features(train_img)
            train_df_t = pd.DataFrame(train_features_t)
            train_data_m = train_df_t.iloc[0].tolist()
            blank = ' '
            train_data = blank.join('%s' % a for a in train_data_m)
            train_features = openpyxl.load_workbook(train_features_path)
            train_features_s = train_features.active
            sheetnames = train_features.get_sheet_names()
            table = train_features.get_sheet_by_name(sheetnames[0])
            table = train_features.active
            nrows = train_features_s.max_row  # 获得行数
            ncolumns = table.max_column  # 获得行数
            values = [train_data]
            for value in values:
                table.cell(nrows + 1, 1).value = value
                nrows = nrows + 1
            train_features.save(train_features_path)